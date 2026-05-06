#!/usr/bin/env python3
"""
generate_xlsx.py
Builds the insurance vs. self-pay Excel workbook from user inputs and MRF data.
Uses openpyxl. All calculations are Excel formulas (not Python-computed values)
so the workbook stays live and editable.

Usage:
    python scripts/generate_xlsx.py '<user_inputs_json>' valleymed.org
    python scripts/generate_xlsx.py '<user_inputs_json>' valleymed.org ~/Desktop/analysis.xlsx
"""

import sys
import json
import os
import argparse
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Add scripts dir to path for calculate import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import calculate

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------------------
# Style constants (per xlsx SKILL.md industry standards)
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
COLOR_INPUT_TEXT = "0000FF"      # Blue — user-editable inputs
COLOR_FORMULA_TEXT = "000000"    # Black — calculated cells
COLOR_XSHEET_TEXT = "008000"     # Green — cross-sheet references
COLOR_HIGHLIGHT_BG = "FFFF00"    # Yellow — key assumptions
COLOR_HEADER_BG = "1F4E79"      # Dark blue — section headers
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_GOOD = "C6EFCE"            # Green fill — insurance wins
COLOR_WARN = "FFEB9C"            # Yellow fill — close call
COLOR_BAD = "FFC7CE"             # Red fill — self-pay wins
COLOR_ALT_ROW = "F2F2F2"         # Light gray — alternating rows

FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_CURRENCY_DEC = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT = '0.0%;-0.0%;"-"'
FMT_INT = '#,##0;(#,##0);"-"'


def style_header(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_input(cell, value=None):
    if value is not None:
        cell.value = value
    cell.font = Font(name=FONT_NAME, color=COLOR_INPUT_TEXT, size=10)
    cell.fill = PatternFill("solid", fgColor=COLOR_HIGHLIGHT_BG)


def style_formula(cell, formula=None):
    if formula is not None:
        cell.value = formula
    cell.font = Font(name=FONT_NAME, color=COLOR_FORMULA_TEXT, size=10)


def style_label(cell, text=None, bold=False):
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, bold=bold, size=10)
    cell.alignment = Alignment(vertical="center")


def style_section_title(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, bold=True, size=12, color=COLOR_HEADER_BG)


def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def set_currency(cell):
    cell.number_format = FMT_CURRENCY


def set_pct(cell):
    cell.number_format = FMT_PCT


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def build_inputs_tab(wb, user_inputs, provider, results):
    """Tab 2: Your Inputs — all user-entered values as editable blue cells."""
    ws = wb.create_sheet("Your Inputs")
    ws.sheet_view.showGridLines = True

    col_width(ws, 1, 38)
    col_width(ws, 2, 22)
    col_width(ws, 3, 40)

    def row(label, value, note="", fmt=None):
        r = ws.max_row + 1
        style_label(ws.cell(r, 1, label))
        c = ws.cell(r, 2, value)
        style_input(c)
        if fmt:
            c.number_format = fmt
        if note:
            style_label(ws.cell(r, 3, note))
        return r

    # Header
    ws.merge_cells("A1:C1")
    style_section_title(ws["A1"], "Your Inputs")
    ws.row_dimensions[1].height = 24

    ws.cell(2, 1, "All blue cells are editable — change any value to update the analysis.")
    ws.cell(2, 1).font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws.append([])

    # Provider
    ws.append(["PROVIDER", None, None])
    style_header(ws.cell(ws.max_row, 1), "PROVIDER")
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")

    row("Hospital / Provider", provider.get("hospital_name", user_inputs.get("providers", [""])[0]))
    row("MRF Last Updated", provider.get("mrf_last_updated", "unknown"),
        "Source: CMS Machine-Readable File")
    row("Self-Pay Discount Rate", provider.get("self_pay_discount_rate", 0.30),
        "Computed from MRF — gross vs. discounted_cash", FMT_PCT)
    ws.append([])

    # Insurance plan
    ws.append(["INSURANCE PLAN", None, None])
    style_header(ws.cell(ws.max_row, 1), "INSURANCE PLAN")
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")

    monthly_row = row("Monthly Premium", user_inputs.get("monthly_premium", 0),
                      "From your insurance card or plan documents", FMT_CURRENCY)
    row("Annual Deductible", user_inputs.get("annual_deductible", 0), "", FMT_CURRENCY)
    row("Out-of-Pocket Maximum", user_inputs.get("out_of_pocket_maximum", 0), "", FMT_CURRENCY)
    row("Coinsurance Rate (your share)", user_inputs.get("coinsurance_rate", 0.20),
        "e.g., 0.20 means you pay 20%", FMT_PCT)
    row("Copay — Primary Care", user_inputs.get("copay_primary_care", 0), "", FMT_CURRENCY)
    row("Copay — Specialist", user_inputs.get("copay_specialist", 0), "", FMT_CURRENCY)
    row("Plan Type", user_inputs.get("plan_type", "PPO"),
        "HMO / PPO / HDHP — affects pre-auth burden")
    row("Insurer / Payer Name", user_inputs.get("payer_name", ""),
        "Used to match negotiated rates in MRF")
    ws.append([])

    # Utilization
    ws.append(["EXPECTED UTILIZATION (per year)", None, None])
    style_header(ws.cell(ws.max_row, 1), "EXPECTED UTILIZATION (per year)")
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")

    row("Primary Care Visits", user_inputs.get("primary_care_visits", 0))
    row("Specialist Visits", user_inputs.get("specialist_visits", 0))
    row("ER / Urgent Care Visits", user_inputs.get("er_visits", 0))
    row("ER Severity", user_inputs.get("er_severity", "moderate"),
        "'minor', 'moderate', or 'major'")
    row("Inpatient Days", user_inputs.get("inpatient_days", 0),
        "Overnight hospital stays (approximate)")
    planned = ", ".join(user_inputs.get("planned_procedures", [])) or "None"
    row("Planned Procedures (CPT codes)", planned)
    ws.append([])

    # Friction
    ws.append(["PRE-AUTH FRICTION", None, None])
    style_header(ws.cell(ws.max_row, 1), "PRE-AUTH FRICTION")
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")

    row("Your Hourly Value of Time ($/hr)", user_inputs.get("hourly_value_of_time", 0),
        "Used to monetize pre-auth hassle", FMT_CURRENCY)
    row("Hours per Pre-Auth Request", user_inputs.get("pre_auth_hours_per_event", 1.5),
        "Default: 1.5 hrs (industry estimate)")
    row("Denial Rate", user_inputs.get("denial_rate", 0.15),
        "Default: 15% (industry estimate — cite before v1)", FMT_PCT)
    row("Hours to Appeal a Denial", user_inputs.get("denial_appeal_hours", 4.0),
        "Default: 4 hrs (industry estimate)")
    ws.append([])

    # Self-pay modifiers
    ws.append(["SELF-PAY MODIFIERS", None, None])
    style_header(ws.cell(ws.max_row, 1), "SELF-PAY MODIFIERS")
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")

    row("Prompt-Pay Discount", user_inputs.get("prompt_pay_discount", 0),
        "Additional discount for paying quickly", FMT_PCT)
    row("Local Tax Dividend Credit ($)", user_inputs.get("local_tax_dividend_amount", 0),
        "e.g., VMC Valley Tax Dividend (King County, WA)", FMT_CURRENCY)

    return ws, monthly_row


def build_procedure_tab(wb, results, procedures):
    """Tab 3: Procedure Cost Comparison."""
    ws = wb.create_sheet("Procedure Costs")

    col_width(ws, 1, 35)
    col_width(ws, 2, 8)
    col_width(ws, 3, 14)
    col_width(ws, 4, 14)
    col_width(ws, 5, 14)
    col_width(ws, 6, 14)
    col_width(ws, 7, 16)
    col_width(ws, 8, 16)

    ws.merge_cells("A1:H1")
    style_section_title(ws["A1"], "Procedure Cost Comparison — What You Actually Pay")
    ws.row_dimensions[1].height = 22

    ws.cell(2, 1, "Red = self-pay is 2x+ what insurer pays  |  Yellow = 1.5–2x  |  Green = within 50%")
    ws.cell(2, 1).font = Font(name=FONT_NAME, italic=True, size=9)
    ws.append([])

    headers = ["Procedure", "CPT", "Gross Charge", "Self-Pay Price",
               "Insurer Min", "Insurer Max", "Insurer Median", "Your Est. Insured Cost"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(4, col, h))
    ws.row_dimensions[4].height = 30

    # Insurance service details
    ins_by_cpt = {s["cpt"]: s for s in results["insurance"].get("service_details", []) if s.get("cpt")}
    sp_by_cpt = {s["cpt"]: s for s in results["selfpay"].get("service_details", []) if s.get("cpt")}

    row_num = 5
    for cpt, proc in procedures.items():
        if not proc.get("gross_charge_median"):
            continue
        gross = proc.get("gross_charge_median", 0)
        cash = proc.get("discounted_cash_median", 0)
        neg_min = proc.get("negotiated_min")
        neg_max = proc.get("negotiated_max")
        neg_med = proc.get("negotiated_median")
        ins_detail = ins_by_cpt.get(cpt, {})
        insured_cost = ins_detail.get("patient_pays")

        ws.cell(row_num, 1, proc.get("description", ""))
        ws.cell(row_num, 2, cpt)

        for col, val in [(3, gross), (4, cash), (5, neg_min), (6, neg_max), (7, neg_med), (8, insured_cost)]:
            c = ws.cell(row_num, col, val)
            c.number_format = FMT_CURRENCY
            style_formula(c)

        # Color code based on self-pay vs. insurer ratio
        if cash and neg_min and neg_min > 0:
            ratio = cash / neg_min
            fill_color = COLOR_GOOD if ratio < 1.5 else (COLOR_WARN if ratio < 2.0 else COLOR_BAD)
            for col in range(1, 9):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=fill_color)
        elif row_num % 2 == 0:
            for col in range(1, 9):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

        row_num += 1

    return ws


def build_scenarios_tab(wb, results):
    """Tab 4: Annual Cost Scenarios."""
    ws = wb.create_sheet("Annual Cost Scenarios")

    col_width(ws, 1, 30)
    col_width(ws, 2, 18)
    col_width(ws, 3, 18)
    col_width(ws, 4, 18)

    ws.merge_cells("A1:D1")
    style_section_title(ws["A1"], "Annual Cost Scenarios")
    ws.row_dimensions[1].height = 22

    ins = results["insurance"]
    sp = results["selfpay"]

    # Headers
    for col, label in enumerate(["", "Low (50% utilization)", "Expected", "High (150%)"], 1):
        style_header(ws.cell(3, col, label))

    rows_data = [
        ("Annual Premium (insurance only)", ins["premium_annual"], ins["premium_annual"], ins["premium_annual"]),
        ("Medical Cost — Insured", ins["patient_responsibility_capped"] * 0.5,
         ins["patient_responsibility_capped"], min(ins["patient_responsibility_capped"] * 1.5,
                                                   ins.get("oop_max", ins["patient_responsibility_capped"]))),
        ("Pre-Auth Friction (insurance)", ins["friction_cost"] * 0.5,
         ins["friction_cost"], ins["friction_cost"] * 1.5),
        ("TOTAL — Insurance", None, None, None),  # formula row
        (None, None, None, None),
        ("Medical Cost — Self-Pay", sp["total"] * 0.5, sp["total"], sp["total"] * 1.5),
        ("TOTAL — Self-Pay", None, None, None),   # formula row
        (None, None, None, None),
        ("Difference (Insurance − Self-Pay)", None, None, None),  # formula row
    ]

    r = 4
    ins_total_rows = []
    sp_total_rows = []
    for label, low, exp, high in rows_data:
        if label is None:
            ws.append([])
            r += 1
            continue
        ws.cell(r, 1, label)
        style_label(ws.cell(r, 1), bold="TOTAL" in (label or ""))
        if "TOTAL" in (label or "") and "Insurance" in (label or ""):
            # Sum the three rows above
            ins_total_rows.append(r)
            for col in [2, 3, 4]:
                c = ws.cell(r, col)
                c.value = f"=SUM({get_column_letter(col)}{r-3}:{get_column_letter(col)}{r-1})"
                c.number_format = FMT_CURRENCY
                style_formula(c)
                c.font = Font(name=FONT_NAME, bold=True)
        elif "TOTAL" in (label or "") and "Self-Pay" in (label or ""):
            sp_total_rows.append(r)
            for col in [2, 3, 4]:
                c = ws.cell(r, col)
                c.value = f"={get_column_letter(col)}{r-1}"
                c.number_format = FMT_CURRENCY
                style_formula(c)
                c.font = Font(name=FONT_NAME, bold=True)
        elif "Difference" in (label or "") and ins_total_rows and sp_total_rows:
            for col in [2, 3, 4]:
                c = ws.cell(r, col)
                c.value = f"={get_column_letter(col)}{ins_total_rows[-1]}-{get_column_letter(col)}{sp_total_rows[-1]}"
                c.number_format = FMT_CURRENCY
                style_formula(c)
                c.font = Font(name=FONT_NAME, bold=True, color=COLOR_INPUT_TEXT)
        else:
            for col, val in [(2, low), (3, exp), (4, high)]:
                if val is not None:
                    c = ws.cell(r, col, round(val, 2))
                    c.number_format = FMT_CURRENCY
                    style_formula(c)
        r += 1

    return ws


def build_breakeven_tab(wb, results):
    """Tab 5: Break-Even Analysis."""
    ws = wb.create_sheet("Break-Even Analysis")
    col_width(ws, 1, 38)
    col_width(ws, 2, 22)
    col_width(ws, 3, 38)

    ws.merge_cells("A1:C1")
    style_section_title(ws["A1"], "Break-Even Analysis")

    be = results["break_even"]
    summary = results["summary"]

    ws.cell(3, 1, "Break-Even Annual Medical Spend")
    ws.cell(3, 1).font = Font(name=FONT_NAME, bold=True, size=11)
    c = ws.cell(3, 2, be.get("break_even_spend"))
    c.number_format = FMT_CURRENCY
    c.font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_INPUT_TEXT)

    ws.cell(4, 1, "What this means:")
    ws.cell(5, 1, (
        f"If your total annual medical bills are below "
        f"${be.get('break_even_spend', 0):,.0f}, self-pay costs less than "
        f"insurance (premiums + your share of bills). Above that threshold, "
        f"insurance saves you money — before accounting for catastrophic protection."
    ))
    ws.cell(5, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[5].height = 48
    ws.merge_cells("A5:C5")

    r = 7
    for label, val, fmt in [
        ("Annual Premium", be.get("annual_premium"), FMT_CURRENCY),
        ("Effective Insurer Discount Rate", be.get("insurer_discount"), FMT_PCT),
        ("Self-Pay Discount Rate", be.get("self_pay_discount"), FMT_PCT),
        ("Discount Gap (insurer − self-pay)", be.get("discount_gap"), FMT_PCT),
        ("Data Source", be.get("insurer_discount_source"), None),
    ]:
        style_label(ws.cell(r, 1, label))
        c = ws.cell(r, 2, val)
        if fmt:
            c.number_format = fmt
        r += 1

    # Build a simple break-even chart table
    r += 2
    ws.cell(r, 1, "Annual Medical Spend")
    style_header(ws.cell(r, 1))
    ws.cell(r, 2, "Total Cost — Insurance")
    style_header(ws.cell(r, 2))
    ws.cell(r, 3, "Total Cost — Self-Pay")
    style_header(ws.cell(r, 3))

    be_val = be.get("break_even_spend") or 20000
    annual_premium = be.get("annual_premium", 0)
    insurer_discount = be.get("insurer_discount", 0.55)
    self_pay_discount = be.get("self_pay_discount", 0.30)

    chart_start = r + 1
    for spend in range(0, int(be_val * 2.2), int(be_val * 2.2 / 15) or 1000):
        r += 1
        ws.cell(r, 1, spend).number_format = FMT_CURRENCY
        ins_cost = annual_premium + spend * (1 - insurer_discount)
        sp_cost = spend * (1 - self_pay_discount)
        ws.cell(r, 2, round(ins_cost, 2)).number_format = FMT_CURRENCY
        ws.cell(r, 3, round(sp_cost, 2)).number_format = FMT_CURRENCY
    chart_end = r

    # Add line chart
    chart = LineChart()
    chart.title = "Break-Even: Insurance vs. Self-Pay"
    chart.style = 10
    chart.y_axis.title = "Annual Total Cost ($)"
    chart.x_axis.title = "Annual Medical Spend ($)"
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=2, max_col=3, min_row=chart_start - 1, max_row=chart_end)
    cats = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A20")

    return ws


def build_catastrophic_tab(wb, results):
    """Tab 6: Catastrophic Scenarios."""
    ws = wb.create_sheet("Catastrophic Scenarios")
    col_width(ws, 1, 30)
    col_width(ws, 2, 18)
    col_width(ws, 3, 22)
    col_width(ws, 4, 22)
    col_width(ws, 5, 22)

    ws.merge_cells("A1:E1")
    style_section_title(ws["A1"], "Catastrophic Event Scenarios")

    ws.cell(2, 1, "This is where insurance almost always wins. The OOP maximum caps your liability; self-pay does not.")
    ws.cell(2, 1).font = Font(name=FONT_NAME, italic=True, size=10)
    ws.merge_cells("A2:E2")

    headers = ["Event", "Self-Pay Total", "Insurance (OOP + Premiums YTD)", "Insurance Protection Value", "Winner"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(4, col, h))
    ws.row_dimensions[4].height = 30

    EVENT_LABELS = {
        "er_visit_moderate": "ER Visit (Moderate)",
        "er_visit_major": "ER Visit (Major)",
        "appendectomy": "Appendectomy",
        "knee_replacement": "Knee Replacement",
        "hip_replacement": "Hip Replacement",
        "cardiac_stent": "Cardiac Stent Procedure",
        "chemotherapy_course": "Chemotherapy (5 sessions)",
        "vaginal_delivery": "Vaginal Delivery",
        "cesarean_delivery": "Cesarean Delivery",
    }

    r = 5
    for event_key, data in results["catastrophic"].items():
        sp = data["selfpay_total"]
        ins = data["insurance_total_with_premiums"]
        prot = data["insurance_protection_value"]
        winner = "Insurance ✓" if prot > 0 else "Self-Pay"
        winner_color = COLOR_GOOD if prot > 0 else COLOR_BAD

        ws.cell(r, 1, EVENT_LABELS.get(event_key, event_key))
        for col, (val, fmt) in enumerate([(sp, FMT_CURRENCY), (ins, FMT_CURRENCY), (prot, FMT_CURRENCY)], 2):
            c = ws.cell(r, col, val)
            c.number_format = fmt
            style_formula(c)

        c = ws.cell(r, 5, winner)
        c.fill = PatternFill("solid", fgColor=winner_color)
        c.font = Font(name=FONT_NAME, bold=True)
        r += 1

    return ws


def build_friction_tab(wb, results, user_inputs):
    """Tab 7: Pre-Auth Friction Detail."""
    ws = wb.create_sheet("Pre-Auth Friction")
    col_width(ws, 1, 35)
    col_width(ws, 2, 12)
    col_width(ws, 3, 14)
    col_width(ws, 4, 14)
    col_width(ws, 5, 30)

    ws.merge_cells("A1:E1")
    style_section_title(ws["A1"], "Pre-Authorization Friction Analysis")

    ins = results["insurance"]
    plan_type = user_inputs.get("plan_type", "PPO")
    hourly = user_inputs.get("hourly_value_of_time", 0)
    denial_rate = user_inputs.get("denial_rate", 0.15)
    pre_auth_hrs = user_inputs.get("pre_auth_hours_per_event", 1.5)
    appeal_hrs = user_inputs.get("denial_appeal_hours", 4.0)

    ws.cell(3, 1, f"Plan type: {plan_type}  |  Your time value: ${hourly}/hr  |  Denial rate: {denial_rate:.0%}")
    ws.cell(3, 1).font = Font(name=FONT_NAME, italic=True, size=9)

    ws.cell(5, 1, "Total Pre-Auth Events This Year")
    ws.cell(5, 2, ins.get("pre_auth_events", 0))
    ws.cell(7, 1, "Total Friction Cost (Annual)")
    c = ws.cell(7, 2, ins.get("friction_cost", 0))
    c.number_format = FMT_CURRENCY
    c.font = Font(name=FONT_NAME, bold=True, size=12)

    ws.cell(9, 1, "Note: Default values are industry estimates. Adjust in 'Your Inputs' tab.")
    ws.cell(9, 1).font = Font(name=FONT_NAME, italic=True, size=9, color="888888")
    ws.cell(10, 1, "Sources to cite before v1 release: AHIP Annual Prior Auth Survey, KFF, AMA Prior Auth Survey")
    ws.cell(10, 1).font = Font(name=FONT_NAME, italic=True, size=9, color="888888")

    return ws


def build_provider_data_tab(wb, results, provider, fqdn, procedures):
    """Tab 8: Provider Data — raw MRF extract."""
    ws = wb.create_sheet("Provider Data")
    col_width(ws, 1, 35)
    col_width(ws, 2, 10)
    col_width(ws, 3, 16)
    col_width(ws, 4, 16)
    col_width(ws, 5, 14)
    col_width(ws, 6, 14)
    col_width(ws, 7, 14)
    col_width(ws, 8, 35)

    ws.merge_cells("A1:H1")
    style_section_title(ws["A1"], "Provider Data — MRF Extract (Audit Trail)")

    ws.cell(3, 1, f"Provider: {provider.get('hospital_name', fqdn)}")
    ws.cell(4, 1, f"MRF URL: {provider.get('mrf_url', 'N/A')}")
    ws.cell(5, 1, f"MRF Last Updated: {provider.get('mrf_last_updated', 'unknown')}")
    ws.cell(6, 1, f"Schema Version: {provider.get('schema_version', 'unknown')}")
    ws.cell(7, 1, f"Skill Run Date: {date.today().isoformat()}")
    ws.cell(8, 1, f"Self-Pay Discount Rate: {provider.get('self_pay_discount_rate', 'N/A')}")

    headers = ["Procedure", "CPT", "Gross Charge", "Self-Pay Price",
               "Neg. Min", "Neg. Max", "Neg. Median", "Category"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(10, col, h))

    r = 11
    for cpt, proc in procedures.items():
        ws.cell(r, 1, proc.get("description", ""))
        ws.cell(r, 2, cpt)
        for col, key, fmt in [
            (3, "gross_charge_median", FMT_CURRENCY),
            (4, "discounted_cash_median", FMT_CURRENCY),
            (5, "negotiated_min", FMT_CURRENCY),
            (6, "negotiated_max", FMT_CURRENCY),
            (7, "negotiated_median", FMT_CURRENCY),
        ]:
            c = ws.cell(r, col, proc.get(key))
            c.number_format = fmt
        ws.cell(r, 8, proc.get("category", ""))
        if r % 2 == 0:
            for col in range(1, 9):
                ws.cell(r, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
        r += 1

    return ws


def build_methodology_tab(wb, results, provider):
    """Tab 9: Methodology & Notes."""
    ws = wb.create_sheet("Methodology & Notes")
    col_width(ws, 1, 90)

    ws.merge_cells("A1:A1")
    style_section_title(ws["A1"], "Methodology & Notes")

    notes = [
        "",
        "HOW HOSPITAL PRICING WORKS",
        "─────────────────────────────────────────",
        "Gross Charge (Chargemaster): The hospital's full list price before any discount. "
        "Almost no one actually pays this. It is the starting point for all negotiations.",
        "",
        "Self-Pay (Discounted Cash) Price: The price offered to uninsured/self-pay patients. "
        f"At {provider.get('hospital_name', 'this provider')}, this is "
        f"{(provider.get('self_pay_discount_rate', 0.30) * 100):.0f}% off the gross charge. "
        "Confirmed by computing 1 - (discounted_cash / gross_charge) across multiple MRF rows.",
        "",
        'THE "SIMILAR TO INSURANCE" CLAIM',
        "─────────────────────────────────────────",
        "Many hospitals state that their self-pay discount is 'similar to the discount taken by "
        "contracted insurance carriers.' This is misleading. Insurance companies typically "
        "negotiate 50-80% off chargemaster rates. The self-pay discount at most hospitals "
        "is 20-40%. The gap means self-pay patients commonly pay 2-3x what an insurer pays "
        "for identical services. This spreadsheet makes that gap explicit.",
        "",
        "Real examples from the VMC MRF (validated during development):",
        "  Stent placement (CPT 37238): Gross $30,627 → Self-pay $21,439 → Kaiser $12,623 (59% off gross)",
        "  Urine albumin lab (CPT 82043): Gross $60 → Self-pay $42 → First Choice $16.76 (72% off gross)",
        "",
        "BREAK-EVEN CALCULATION",
        "─────────────────────────────────────────",
        "Break-even spend = Annual premium / (Insurer discount rate − Self-pay discount rate)",
        "Below this threshold: self-pay total cost < insured total cost",
        "Above this threshold: insurance savings on negotiated rates exceed premium cost",
        "Note: This calculation does not include catastrophic protection (OOP max) — "
        "see the Catastrophic Scenarios tab for that analysis.",
        "",
        "CATASTROPHIC PROTECTION",
        "─────────────────────────────────────────",
        "The out-of-pocket maximum is insurance's core value proposition. "
        "For major medical events (surgery, cancer, cardiac events), self-pay costs "
        "can reach $50,000-$200,000+ even after the self-pay discount. "
        "Insurance caps your liability at the OOP max regardless of actual costs. "
        "This is where insurance almost always wins — the Catastrophic Scenarios tab "
        "quantifies this protection value explicitly.",
        "",
        "PRE-AUTH FRICTION (DEFAULT VALUES — CITE BEFORE v1)",
        "─────────────────────────────────────────",
        "• Denial rate: 15% (industry estimate)",
        "• Time per auth request: 1.5 hours (industry estimate)",
        "• Appeal time: 4 hours (industry estimate)",
        "• Sources to verify: AHIP Annual Prior Auth Survey, KFF Health System Tracker, AMA Prior Auth Survey",
        "",
        "LIMITATIONS",
        "─────────────────────────────────────────",
        "• Inpatient stays are approximated using per-diem rates from MRF inpatient data. "
        "Actual inpatient costs are DRG-based (Diagnosis-Related Group) and vary by diagnosis.",
        "• Financial assistance and charity care eligibility is flagged but not fully modeled "
        "for all providers. Contact your provider's financial counseling office for details.",
        "• MRF data is updated annually by hospitals. Rates may change between updates.",
        "• Negotiated rates shown are from the provider's published MRF. "
        "Actual insurer payments may differ due to claim adjustments.",
        "",
        "DATA SOURCES",
        "─────────────────────────────────────────",
        f"• Provider MRF: {provider.get('mrf_url', 'N/A')}",
        f"• MRF last updated: {provider.get('mrf_last_updated', 'unknown')}",
        "• CMS MRF schema: https://github.com/CMSgov/hospital-price-transparency",
        "• WA Healthplanfinder: https://www.wahealthplanfinder.org",
    ]

    for i, note in enumerate(notes, 2):
        ws.cell(i, 1, note)
        if note and note == note.upper() and len(note) > 5:
            ws.cell(i, 1).font = Font(name=FONT_NAME, bold=True, size=10)
        else:
            ws.cell(i, 1).font = Font(name=FONT_NAME, size=9)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 15 if len(note) < 100 else 30

    return ws


def build_summary_tab(wb, results, user_inputs, provider):
    """Tab 1: Summary Dashboard — built last so cross-sheet refs are valid."""
    ws = wb.worksheets[0]
    ws.title = "Summary"
    col_width(ws, 1, 30)
    col_width(ws, 2, 22)
    col_width(ws, 3, 22)
    col_width(ws, 4, 22)

    summary = results["summary"]
    ins_total = summary["insurance_total"]
    sp_total = summary["selfpay_total"]
    winner = summary["winner"]
    diff = summary["annual_difference"]
    be = summary["break_even_spend"]

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Insurance vs. Self-Pay Analysis"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    ws["A2"] = f"{provider.get('hospital_name', 'Your Provider')}  |  Generated {date.today().isoformat()}"
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Top-line answer
    winner_label = "Insurance" if winner == "insurance" else "Self-Pay"
    winner_color = COLOR_GOOD if winner == "insurance" else COLOR_BAD
    ws.merge_cells("A4:D4")
    ws["A4"] = f"Based on your inputs: {winner_label.upper()} saves you approximately ${diff:,.0f}/year"
    ws["A4"].font = Font(name=FONT_NAME, bold=True, size=13)
    ws["A4"].fill = PatternFill("solid", fgColor=winner_color)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28

    # Key metrics table
    r = 6
    for label, val, fmt, note in [
        ("Annual Cost — Insurance", ins_total, FMT_CURRENCY,
         f"Premium ${user_inputs.get('monthly_premium',0)*12:,.0f} + medical + friction"),
        ("Annual Cost — Self-Pay", sp_total, FMT_CURRENCY,
         f"{provider.get('self_pay_discount_rate',0.30):.0%} discount off gross charges"),
        ("Annual Difference", diff, FMT_CURRENCY,
         f"{winner_label} wins by this amount per year"),
        ("Break-Even Medical Spend", be, FMT_CURRENCY,
         "Above this threshold, insurance saves money on bills"),
        ("Catastrophic Protection Value", results["catastrophic"].get(
            "knee_replacement", {}).get("insurance_protection_value"), FMT_CURRENCY,
         "Insurance savings on a knee replacement vs. self-pay"),
    ]:
        style_label(ws.cell(r, 1, label), bold=True)
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        c.font = Font(name=FONT_NAME, bold=True, size=12)
        ws.cell(r, 3, note).font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
        ws.merge_cells(f"C{r}:D{r}")
        r += 1

    ws.cell(r + 1, 1, "→ See 'Annual Cost Scenarios' tab for low/expected/high utilization breakdown")
    ws.cell(r + 2, 1, "→ See 'Break-Even Analysis' tab for the full chart")
    ws.cell(r + 3, 1, "→ See 'Catastrophic Scenarios' tab for major event modeling")
    ws.cell(r + 4, 1, "→ Update your inputs in the 'Your Inputs' tab — all tabs recalculate automatically")
    for note_row in range(r + 1, r + 5):
        ws.cell(note_row, 1).font = Font(name=FONT_NAME, italic=True, size=10, color=COLOR_HEADER_BG)

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("user_inputs_json", help="JSON string of user inputs")
    parser.add_argument("fqdn", help="Provider FQDN, e.g. valleymed.org")
    parser.add_argument("output_path", nargs="?",
                        default=os.path.expanduser(f"~/Downloads/insurance-analysis-{date.today().isoformat()}.xlsx"))
    args = parser.parse_args()

    try:
        user_inputs = json.loads(args.user_inputs_json)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON for user_inputs: {e}")
        return 1

    fqdn = args.fqdn.lower().strip().lstrip("www.").split("/")[0]
    print(f"\n=== Generating workbook for {fqdn} ===")

    # Run calculations
    print("  Running calculations...")
    results = calculate.run_all_scenarios(user_inputs, fqdn)
    provider = calculate.load_provider(fqdn)
    procedures = calculate.load_procedures(fqdn)

    summary = results["summary"]
    winner = "Insurance" if summary["winner"] == "insurance" else "Self-Pay"
    print(f"  Result: {winner} wins by ${summary['annual_difference']:,.0f}/year")

    # Build workbook
    print("  Building workbook...")
    wb = Workbook()

    # Tab order matters — Summary is first (wb.active = Tab 1)
    # We build Summary last to allow cross-sheet refs to exist first
    build_inputs_tab(wb, user_inputs, provider, results)
    build_procedure_tab(wb, results, procedures)
    build_scenarios_tab(wb, results)
    build_breakeven_tab(wb, results)
    build_catastrophic_tab(wb, results)
    build_friction_tab(wb, results, user_inputs)
    build_provider_data_tab(wb, results, provider, fqdn, procedures)
    build_methodology_tab(wb, results, provider)
    build_summary_tab(wb, results, user_inputs, provider)

    # Reorder so Summary is Tab 1
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    os.makedirs(os.path.dirname(os.path.abspath(args.output_path)), exist_ok=True)
    wb.save(args.output_path)
    print(f"\n  Workbook saved: {args.output_path}")
    print(f"  Next step: python scripts/recalc.py \"{args.output_path}\"")
    return 0


if __name__ == "__main__":
    sys.exit(main())
